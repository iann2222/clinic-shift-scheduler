"""Openpyxl adapter for human-readable formal schedule workbooks."""

from __future__ import annotations

import os
import tempfile
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..enums import EmploymentType, FullTimeClass, Period
from ..models import NormalizedScheduleInput
from ..optimization import OptimizationStageStatus
from ..output import FormalScheduleOutput, RatioValue, ScheduleCellKind
from .files import (
    DEFAULT_OUTPUT_DIRECTORY,
    build_output_paths,
    build_provisional_output_paths,
    prepare_target,
    require_formal_result,
    require_provisional_result,
    schedule_month,
)


WORKSHEET_NAMES = (
    "月班表",
    "個人班型摘要",
    "個人詳細統計",
    "類別與公平性統計",
    "求解與驗證資訊",
)
PROVISIONAL_WORKSHEET_NAMES = ("暫存結果說明",) + WORKSHEET_NAMES

_DARK_BLUE = "1F4E78"
_HEADER_BLUE = "C5DFF0"
_MORNING = "FCE8A3"
_AFTERNOON = "C9DFF0"
_EVENING = "D7CEE6"
_CLOSED = "C9C9C9"
_WEEKEND = "F4CEB8"
_WHITE = "FFFFFF"
_GRID = Side(style="thin", color="7890A0")
_BORDER = Border(left=_GRID, right=_GRID, top=_GRID, bottom=_GRID)

_PERIOD_LABELS = {
    Period.MORNING: "早",
    Period.AFTERNOON: "午",
    Period.EVENING: "晚",
}
_PERIOD_FILLS = {
    Period.MORNING: _MORNING,
    Period.AFTERNOON: _AFTERNOON,
    Period.EVENING: _EVENING,
}
_ROLE_LABELS = {
    "reception": "櫃台",
    "nursing": "跟診",
    "assistant": "跟診",
}
_WEEKDAY_LABELS = ("一", "二", "三", "四", "五", "六", "日")
_EMPLOYEE_FONT_COLORS = (
    "2860AD",
    "AD2833",
    "28AD60",
    "8128AD",
    "AD8128",
    "28A2AD",
    "8CAE29",
    "3FAE29",
    "3F29AE",
    "AE2986",
)


def _header(cell: Cell, *, fill: str = _DARK_BLUE) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=_WHITE if fill == _DARK_BLUE else "000000")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER


def _style_table(
    sheet: Worksheet,
    *,
    header_row: int,
    first_data_row: int,
    last_row: int,
    last_column: int,
) -> None:
    for cell in sheet[header_row]:
        if cell.column <= last_column:
            _header(cell)
    for row in sheet.iter_rows(
        min_row=first_data_row,
        max_row=last_row,
        min_col=1,
        max_col=last_column,
    ):
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(last_column)}{last_row}"
    )
    sheet.freeze_panes = f"A{first_data_row}"
    sheet.sheet_view.showGridLines = False


def _translated_role(role: str) -> str:
    return _ROLE_LABELS.get(role, role)


def _employee_color_map(output: FormalScheduleOutput) -> dict[str, str]:
    statistics = sorted(
        output.individual_statistics,
        key=lambda item: item.employee_id,
    )
    return {
        item.employee_id: _EMPLOYEE_FONT_COLORS[
            index % len(_EMPLOYEE_FONT_COLORS)
        ]
        for index, item in enumerate(statistics)
    }


def _plain_schedule_header(cell: Cell) -> None:
    cell.fill = PatternFill(fill_type=None)
    cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER


def _build_schedule_sheet(
    workbook: Workbook,
    output: FormalScheduleOutput,
) -> None:
    schedule = output.monthly_schedule
    assert schedule is not None
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "月班表"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "C3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:2"
    sheet.print_options.horizontalCentered = True

    for row_number, label in ((1, "日期"), (2, "星期")):
        for column in (1, 2):
            _plain_schedule_header(sheet.cell(row_number, column))
        sheet.cell(row_number, 1, label)
        sheet.merge_cells(
            start_row=row_number,
            start_column=1,
            end_row=row_number,
            end_column=2,
        )
    sheet.column_dimensions["A"].width = 3
    sheet.column_dimensions["B"].width = 5
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 22

    for column, day in enumerate(schedule.dates, start=3):
        date_cell = sheet.cell(1, column, day)
        date_cell.number_format = "m/d"
        weekday_cell = sheet.cell(2, column, _WEEKDAY_LABELS[day.weekday()])
        header_fill = _WEEKEND if day.weekday() >= 5 else _HEADER_BLUE
        _header(date_cell, fill=header_fill)
        _header(weekday_cell, fill=header_fill)
        sheet.column_dimensions[get_column_letter(column)].width = 11

    row_lookup = {
        (row.period, _translated_role(row.role), row.position): row
        for row in schedule.rows
    }
    employee_colors = _employee_color_map(output)
    role_rows = (
        ("櫃台", 1, "櫃台"),
        ("跟診", 1, "一診"),
        ("跟診", 2, "二診"),
    )
    for period_index, period in enumerate(Period):
        first_row = 3 + period_index * len(role_rows)
        last_row = first_row + len(role_rows) - 1
        fill = PatternFill("solid", fgColor=_PERIOD_FILLS[period])
        for row_number in range(first_row, last_row + 1):
            for column in (1, 2):
                label_cell = sheet.cell(row_number, column)
                label_cell.fill = fill
                label_cell.font = Font(bold=True)
                label_cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                label_cell.border = _BORDER
            sheet.row_dimensions[row_number].height = 27
        sheet.cell(first_row, 1, _PERIOD_LABELS[period])
        sheet.merge_cells(
            start_row=first_row,
            start_column=1,
            end_row=last_row,
            end_column=1,
        )

        reference_row = next(row for row in schedule.rows if row.period is period)
        for offset, (role_label, position, display_label) in enumerate(role_rows):
            row_number = first_row + offset
            sheet.cell(row_number, 2, display_label)
            schedule_row = row_lookup.get((period, role_label, position))
            for date_index, _day in enumerate(schedule.dates):
                if schedule_row is None:
                    reference = reference_row.cells[date_index]
                    source_cell = None
                    if reference.kind is ScheduleCellKind.CLOSED:
                        display = "休診"
                        kind = ScheduleCellKind.CLOSED
                    else:
                        display = "—"
                        kind = ScheduleCellKind.ZERO_DEMAND
                else:
                    source_cell = schedule_row.cells[date_index]
                    display = source_cell.display
                    kind = source_cell.kind

                cell = sheet.cell(row_number, date_index + 3, display)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = _BORDER
                if kind is ScheduleCellKind.CLOSED:
                    cell.fill = PatternFill("solid", fgColor=_CLOSED)
                    cell.font = Font(color="4A4A4A", italic=True, bold=True)
                elif kind is ScheduleCellKind.ZERO_DEMAND:
                    cell.font = Font(color="7A7A7A", bold=True)
                elif source_cell is not None and source_cell.employee_id is not None:
                    cell.font = Font(
                        color=employee_colors[source_cell.employee_id],
                        bold=True,
                    )


def _employment_label(value: EmploymentType) -> str:
    return "正職" if value is EmploymentType.FULL_TIME else "兼職"


def _full_time_class_label(value: FullTimeClass | None) -> str:
    return value.value if value is not None else "N/A"


def _ratio_cell_value(ratio: RatioValue | None) -> float | str:
    return "N/A" if ratio is None or ratio.value is None else ratio.value


def _employee_category_label(
    employment_type: EmploymentType,
    full_time_class: FullTimeClass | None,
) -> str:
    if employment_type is EmploymentType.PART_TIME:
        return "兼職"
    return f"{_full_time_class_label(full_time_class)} 類正職"


def _pattern_summary(ratio: RatioValue) -> str:
    percentage = "N/A" if ratio.value is None else f"{ratio.value:.1%}"
    return f"{ratio.numerator} / {ratio.denominator}（{percentage}）"


def _build_individual_summary_sheet(
    workbook: Workbook,
    output: FormalScheduleOutput,
) -> None:
    sheet = workbook.create_sheet("個人班型摘要")
    employee_colors = _employee_color_map(output)
    headers = (
        "姓名",
        "類別",
        "總班次",
        "出勤日",
        "連續雙班日／出勤日",
        "單節日／出勤日",
        "早＋晚拆班日／出勤日",
        "三節班日／出勤日",
        "週日節數",
        "週日出勤天數",
    )
    sheet.append(headers)
    for stats in output.individual_statistics:
        sheet.append(
            (
                stats.name,
                _employee_category_label(
                    stats.employment_type,
                    stats.full_time_class,
                ),
                stats.total_shifts,
                stats.attendance_days,
                _pattern_summary(
                    stats.ratios["consecutive_double_days"],
                ),
                _pattern_summary(
                    stats.ratios["single_shift_days"],
                ),
                _pattern_summary(
                    stats.ratios["morning_evening_days"],
                ),
                _pattern_summary(
                    stats.ratios["triple_days"],
                ),
                stats.sunday_shifts,
                stats.sunday_attendance_days,
            )
        )
        sheet.cell(sheet.max_row, 1).font = Font(
            color=employee_colors[stats.employee_id],
            bold=True,
        )
    _style_table(
        sheet,
        header_row=1,
        first_data_row=2,
        last_row=sheet.max_row,
        last_column=sheet.max_column,
    )
    widths = (16, 14, 11, 11, 24, 22, 26, 22, 12, 15)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 25
        for column in range(3, sheet.max_column + 1):
            sheet.cell(row, column).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )


def _build_individual_detail_sheet(
    workbook: Workbook,
    output: FormalScheduleOutput,
) -> None:
    sheet = workbook.create_sheet("個人詳細統計")
    employee_colors = _employee_color_map(output)
    overall = output.overall_statistics
    assert overall is not None
    roles = tuple(overall.role_counts)
    ratio_names = tuple(
        sorted(
            {
                name
                for stats in output.individual_statistics
                for name in stats.ratios
            }
        )
    )
    headers = (
        "employee_id",
        "姓名",
        "人員類型",
        "正職類別",
        "fairness_group",
        "shift_mode",
        "目標班次",
        "目標偏差",
        "相對目標偏差 (bp)",
        "總班次",
        *(f"{_translated_role(role)}班次" for role in roles),
        "早班",
        "午班",
        "晚班",
        "出勤日",
        "單節日",
        "連續雙班",
        "早午雙班",
        "午晚雙班",
        "早晚雙班",
        "三節班",
        "星期日班次",
        "假日班次",
        "請假時段",
        "可排時段",
        *(f"{name} 比例" for name in ratio_names),
    )
    sheet.append(headers)
    for stats in output.individual_statistics:
        sheet.append(
            (
                stats.employee_id,
                stats.name,
                _employment_label(stats.employment_type),
                _full_time_class_label(stats.full_time_class),
                stats.fairness_group,
                stats.shift_mode,
                stats.target_shifts,
                stats.target_deviation,
                stats.target_relative_deviation_basis_points,
                stats.total_shifts,
                *(stats.role_counts[role] for role in roles),
                stats.period_counts["morning"],
                stats.period_counts["afternoon"],
                stats.period_counts["evening"],
                stats.attendance_days,
                stats.single_shift_days,
                stats.consecutive_double_days,
                stats.morning_afternoon_days,
                stats.afternoon_evening_days,
                stats.morning_evening_days,
                stats.triple_days,
                stats.sunday_shifts,
                stats.holiday_shifts,
                stats.leave_periods,
                stats.available_periods,
                *(_ratio_cell_value(stats.ratios.get(name)) for name in ratio_names),
            )
        )
        sheet.cell(sheet.max_row, 2).font = Font(
            color=employee_colors[stats.employee_id]
        )
    _style_table(
        sheet,
        header_row=1,
        first_data_row=2,
        last_row=sheet.max_row,
        last_column=sheet.max_column,
    )
    for column in range(1, sheet.max_column + 1):
        header = sheet.cell(1, column).value
        sheet.column_dimensions[get_column_letter(column)].width = (
            18 if column <= 6 else 13
        )
        if isinstance(header, str) and header.endswith("比例"):
            for row in range(2, sheet.max_row + 1):
                if isinstance(sheet.cell(row, column).value, float):
                    sheet.cell(row, column).number_format = "0.0%"


def _build_group_sheet(
    workbook: Workbook,
    output: FormalScheduleOutput,
) -> None:
    sheet = workbook.create_sheet("類別與公平性統計")
    overall = output.overall_statistics
    assert overall is not None
    sheet.append(("A／B 類別偏好理想值與公平退讓",))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    _header(sheet.cell(1, 1))
    sheet.append(
        (
            "正職類別",
            "需求順位",
            "指標",
            "方向",
            "實際值",
            "鎖定實際值",
            "單類理想值",
            "退讓日數",
            "機會日數",
            "退讓比例 (bp)",
        )
    )
    for stats in output.class_preference_statistics:
        sheet.append(
            (
                _full_time_class_label(stats.full_time_class),
                stats.rank.value,
                stats.metric,
                stats.direction.value,
                stats.actual_value,
                stats.locked_actual_value,
                stats.ideal_value,
                stats.regret_days,
                stats.opportunity_days,
                "N/A"
                if stats.regret_basis_points is None
                else stats.regret_basis_points,
            )
        )
    preference_last = sheet.max_row
    _style_table(
        sheet,
        header_row=2,
        first_data_row=3,
        last_row=preference_last,
        last_column=10,
    )

    category_title_row = sheet.max_row + 2
    sheet.cell(category_title_row, 1, "類別統計")
    sheet.merge_cells(
        start_row=category_title_row,
        start_column=1,
        end_row=category_title_row,
        end_column=7,
    )
    _header(sheet.cell(category_title_row, 1))
    category_header_row = category_title_row + 1
    for column, value in enumerate(
        (
            "類別",
            "人員",
            "總班次",
            "連續雙班",
            "單節日",
            "早晚雙班",
            "三節班",
        ),
        start=1,
    ):
        sheet.cell(category_header_row, column, value)
    for category in output.category_statistics:
        sheet.append(
            (
                category.category,
                ", ".join(category.employee_ids),
                category.total_shifts,
                category.consecutive_double_days,
                category.single_shift_days,
                category.morning_evening_days,
                category.triple_days,
            )
        )
    category_last = sheet.max_row
    _style_table(
        sheet,
        header_row=category_header_row,
        first_data_row=category_header_row + 1,
        last_row=category_last,
        last_column=7,
    )

    group_title_row = sheet.max_row + 2
    sheet.cell(group_title_row, 1, "Fairness group 統計")
    sheet.merge_cells(
        start_row=group_title_row,
        start_column=1,
        end_row=group_title_row,
        end_column=6,
    )
    _header(sheet.cell(group_title_row, 1))
    group_header_row = group_title_row + 1
    group_headers = (
        "fairness_group",
        "人員類型",
        "正職類別",
        "指標",
        "各人數值",
        "最大差距",
    )
    for column, value in enumerate(group_headers, start=1):
        sheet.cell(group_header_row, column, value)
    for group in output.fairness_group_statistics:
        for metric, values in group.metric_values.items():
            sheet.append(
                (
                    group.fairness_group,
                    _employment_label(group.employment_type),
                    _full_time_class_label(group.full_time_class),
                    metric,
                    "; ".join(
                        f"{employee_id}={value}"
                        for employee_id, value in values.items()
                    ),
                    group.gaps[metric],
                )
            )
        for metric, values in group.ratio_basis_points.items():
            sheet.append(
                (
                    group.fairness_group,
                    _employment_label(group.employment_type),
                    _full_time_class_label(group.full_time_class),
                    f"{metric} 比例 (bp)",
                    "; ".join(
                        f"{employee_id}="
                        + ("N/A" if value is None else str(value))
                        for employee_id, value in values.items()
                    ),
                    group.ratio_gaps_basis_points[metric],
                )
            )
    _style_table(
        sheet,
        header_row=group_header_row,
        first_data_row=group_header_row + 1,
        last_row=sheet.max_row,
        last_column=6,
    )
    sheet.freeze_panes = "A3"
    widths = (22, 24, 14, 28, 48, 16, 18)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_key_values(
    sheet: Worksheet,
    start_row: int,
    title: str,
    values: Iterable[tuple[str, Any]],
) -> int:
    sheet.cell(start_row, 1, title)
    sheet.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row, end_column=2
    )
    _header(sheet.cell(start_row, 1))
    row = start_row + 1
    for label, value in values:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        for column in (1, 2):
            cell = sheet.cell(row, column)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(row, 1).font = Font(bold=True)
        row += 1
    return row


def _build_solver_sheet(
    workbook: Workbook,
    data: NormalizedScheduleInput,
    output: FormalScheduleOutput,
) -> None:
    sheet = workbook.create_sheet("求解與驗證資訊")
    report = output.validation_report
    overall = output.overall_statistics
    assert report is not None and overall is not None
    formal_values: list[tuple[str, Any]] = [
        ("月份", schedule_month(data)),
        ("輸入 Schema", data.source.schema_version),
        ("正式狀態", output.status.value),
        ("Validation", report.status.value),
        ("總需求", overall.total_demand),
        ("總安排", overall.total_assignments),
        ("未填需求", overall.unfilled_shifts),
        (
            "已完成目標前綴最佳",
            overall.implemented_objective_prefix_optimal,
        ),
    ]
    timing = output.execution_timing
    if timing is not None:
        formal_values.extend(
            (
                ("輸入讀取（秒）", timing.input_loading_seconds),
                (
                    "驗證與正規化（秒）",
                    timing.validation_normalization_seconds,
                ),
                ("前置可行性檢查（秒）", timing.precheck_seconds),
                ("CP-SAT 最佳化（秒）", timing.optimization_seconds),
                (
                    "獨立驗證與結果建立（秒）",
                    timing.result_validation_and_build_seconds,
                ),
                ("排班管線總時間（秒）", timing.scheduling_pipeline_seconds),
                (
                    "首次合法班表（秒）",
                    timing.time_to_first_feasible_schedule,
                ),
                (
                    "正式最佳值證明（秒）",
                    timing.time_to_proven_formal_optimum,
                ),
            )
        )
    telemetry = output.optimization_telemetry
    if telemetry is not None:
        formal_values.extend(
            (
                ("月份日數", telemetry.days),
                ("員工人數", telemetry.employees),
                ("正職人數", telemetry.full_time_employees),
                ("兼職人數", telemetry.part_time_employees),
                ("assignment 變數數", telemetry.assignment_variables),
                ("可排 assignment 比例", telemetry.availability_ratio),
                ("需求節數", telemetry.demand_units),
            )
        )
    row = _write_key_values(
        sheet,
        1,
        "正式結果",
        tuple(formal_values),
    )
    row += 1
    row = _write_key_values(
        sheet,
        row,
        "Objective vector",
        tuple(overall.objective_vector.items()),
    )
    row += 1
    sheet.cell(row, 1, "類別偏好理想值 benchmark")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    _header(sheet.cell(row, 1))
    row += 1
    benchmark_header = row
    benchmark_headers = (
        "正職類別",
        "順位",
        "指標",
        "方向",
        "狀態",
        "理想值",
        "鎖定實際值",
        "機會日數",
        "solver status",
        "wall time (秒)",
    )
    for column, value in enumerate(benchmark_headers, start=1):
        sheet.cell(row, column, value)
    for benchmark in output.preference_benchmarks:
        row += 1
        values = (
            benchmark.full_time_class.value,
            benchmark.rank.value,
            benchmark.metric.value,
            benchmark.direction.value,
            benchmark.status.value,
            benchmark.ideal_value,
            benchmark.locked_actual_value,
            benchmark.opportunity_days,
            benchmark.raw_solver_status,
            benchmark.wall_time_seconds,
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
    _style_table(
        sheet,
        header_row=benchmark_header,
        first_data_row=benchmark_header + 1,
        last_row=row,
        last_column=10,
    )
    row += 2
    sheet.cell(row, 1, "剩餘班型類別鎖")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _header(sheet.cell(row, 1))
    row += 1
    pattern_lock_header = row
    for column, value in enumerate(
        ("正職類別", "指標", "鎖定實際值"), start=1
    ):
        sheet.cell(row, column, value)
    for item in output.class_pattern_locks:
        row += 1
        for column, value in enumerate(
            (
                item.full_time_class.value,
                item.metric.value,
                item.locked_value,
            ),
            start=1,
        ):
            sheet.cell(row, column, value)
    _style_table(
        sheet,
        header_row=pattern_lock_header,
        first_data_row=pattern_lock_header + 1,
        last_row=row,
        last_column=3,
    )
    row += 2
    sheet.cell(row, 1, "最佳化階段")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    _header(sheet.cell(row, 1))
    row += 1
    stage_header = row
    stage_headers = (
        "順序",
        "stage",
        "方向",
        "狀態",
        "objective value",
        "已鎖定",
        "常數證明",
        "solver status",
        "wall time (秒)",
        "best objective bound",
        "conflicts",
        "branches",
    )
    for column, value in enumerate(stage_headers, start=1):
        sheet.cell(row, column, value)
    for index, stage in enumerate(output.optimization_stages, start=1):
        row += 1
        values = (
            index,
            stage.stage.value,
            stage.direction.value,
            stage.status.value,
            stage.objective_value,
            stage.locked,
            stage.constant_proof.value if stage.constant_proof else "",
            stage.raw_solver_status,
            stage.wall_time_seconds,
            stage.best_objective_bound,
            stage.num_conflicts,
            stage.num_branches,
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
        if stage.status is OptimizationStageStatus.OPTIMAL:
            sheet.cell(row, 4).fill = PatternFill("solid", fgColor="E2F0D9")
    _style_table(
        sheet,
        header_row=stage_header,
        first_data_row=stage_header + 1,
        last_row=row,
        last_column=12,
    )
    row += 2
    validation_title = row
    sheet.cell(row, 1, "Validation checks")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _header(sheet.cell(row, 1))
    row += 1
    validation_header = row
    for column, value in enumerate(("檢查項目", "結果", "問題數"), start=1):
        sheet.cell(row, column, value)
    issue_counts: dict[str, int] = {}
    for issue in report.issues:
        issue_counts[issue.category] = issue_counts.get(issue.category, 0) + 1
    for check, passed in report.checks.items():
        row += 1
        sheet.cell(row, 1, check)
        sheet.cell(row, 2, "PASS" if passed else "FAIL")
        sheet.cell(row, 3, issue_counts.get(check, 0))
    _style_table(
        sheet,
        header_row=validation_header,
        first_data_row=validation_header + 1,
        last_row=row,
        last_column=3,
    )
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 28
    for column in range(3, 10):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    sheet.sheet_view.showGridLines = False


def build_workbook(
    data: NormalizedScheduleInput,
    output: FormalScheduleOutput,
) -> Workbook:
    """Build a workbook strictly from an already finalized output model."""

    require_formal_result(output)
    workbook = Workbook()
    workbook.properties.title = f"{schedule_month(data)} 診所排班結果"
    workbook.properties.subject = "Clinic shift scheduler formal result"
    workbook.properties.creator = "clinic-shift-scheduler"
    workbook.properties.description = (
        "Generated from an independently validated formal result model."
    )
    _build_schedule_sheet(workbook, output)
    _build_individual_summary_sheet(workbook, output)
    _build_individual_detail_sheet(workbook, output)
    _build_group_sheet(workbook, output)
    _build_solver_sheet(workbook, data, output)
    return workbook


def build_provisional_workbook(
    data: NormalizedScheduleInput,
    output: FormalScheduleOutput,
) -> Workbook:
    """Build a visibly provisional workbook from a validated FEASIBLE result."""

    require_provisional_result(output)
    workbook = Workbook()
    workbook.properties.title = f"{schedule_month(data)} 目前最佳合法班表"
    workbook.properties.subject = "Unfinished clinic schedule optimization"
    workbook.properties.creator = "clinic-shift-scheduler"
    workbook.properties.description = (
        "Validated FEASIBLE schedule; formal optimization is incomplete."
    )
    _build_schedule_sheet(workbook, output)
    _build_individual_summary_sheet(workbook, output)
    _build_individual_detail_sheet(workbook, output)
    _build_group_sheet(workbook, output)
    _build_solver_sheet(workbook, data, output)

    notice = workbook.create_sheet("暫存結果說明", 0)
    notice.sheet_view.showGridLines = False
    notice.merge_cells("A1:F2")
    notice["A1"] = "目前最佳合法班表"
    notice["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    notice["A1"].fill = PatternFill("solid", fgColor="B45309")
    notice["A1"].alignment = Alignment(
        horizontal="center", vertical="center"
    )
    notice.merge_cells("A4:F6")
    notice["A4"] = (
        "此班表已通過全部硬性規則驗證，但尚未完成全部最佳化，"
        "不代表正式最佳結果。"
    )
    notice["A4"].font = Font(size=14, bold=True, color="7C2D12")
    notice["A4"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    notice["A8"] = "結果狀態"
    notice["B8"] = output.status.value
    notice["A9"] = "Validation"
    notice["B9"] = output.validation_report.status.value
    notice["A10"] = "已完成正式階段"
    notice["B10"] = len(output.optimization_stages)
    notice.column_dimensions["A"].width = 22
    notice.column_dimensions["B"].width = 28
    for column in ("C", "D", "E", "F"):
        notice.column_dimensions[column].width = 14
    notice.freeze_panes = "A8"
    return workbook


def _validate_saved_workbook(path: Path, *, provisional: bool = False) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        expected_names = (
            PROVISIONAL_WORKSHEET_NAMES if provisional else WORKSHEET_NAMES
        )
        if tuple(workbook.sheetnames) != expected_names:
            raise ValueError("exported workbook has an unexpected worksheet structure")
        schedule = workbook["月班表"]
        if schedule["A1"].value != "日期" or schedule["A2"].value != "星期":
            raise ValueError("exported workbook has an invalid monthly schedule header")
    finally:
        workbook.close()


def export_result_excel(
    data: NormalizedScheduleInput,
    output: FormalScheduleOutput,
    *,
    output_directory: str | Path = DEFAULT_OUTPUT_DIRECTORY,
    overwrite: bool = False,
    filename_stem: str | None = None,
) -> Path:
    """Atomically persist and reopen-validate the formal Excel workbook."""

    workbook = build_workbook(data, output)
    target = build_output_paths(
        data,
        output_directory,
        stem=filename_stem,
    ).excel
    prepare_target(target, overwrite=overwrite)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{target.stem}.",
            suffix=".xlsx",
            dir=target.parent,
            delete=False,
        ) as stream:
            temporary = Path(stream.name)
        workbook.save(temporary)
        workbook.close()
        _validate_saved_workbook(temporary)
        os.replace(temporary, target)
    finally:
        workbook.close()
        if temporary is not None and temporary.exists():
            temporary.unlink()
    return target


def export_provisional_result_excel(
    data: NormalizedScheduleInput,
    output: FormalScheduleOutput,
    *,
    output_directory: str | Path = DEFAULT_OUTPUT_DIRECTORY,
    overwrite: bool = False,
) -> Path:
    workbook = build_provisional_workbook(data, output)
    target = build_provisional_output_paths(data, output_directory).excel
    prepare_target(target, overwrite=overwrite)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{target.stem}.",
            suffix=".xlsx",
            dir=target.parent,
            delete=False,
        ) as stream:
            temporary = Path(stream.name)
        workbook.save(temporary)
        workbook.close()
        _validate_saved_workbook(temporary, provisional=True)
        os.replace(temporary, target)
    finally:
        workbook.close()
        if temporary is not None and temporary.exists():
            temporary.unlink()
    return target
