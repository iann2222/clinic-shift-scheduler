"""Printable monthly-schedule PDF derived exclusively from the formal Excel file."""

from __future__ import annotations

import os
import tempfile
from datetime import date, datetime
from importlib.resources import as_file, files
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .excel_exporter import PROVISIONAL_WORKSHEET_NAMES, WORKSHEET_NAMES
from .files import FormalExportError, prepare_target


_SCHEDULE_SHEET = "月班表"
_INDIVIDUAL_SUMMARY_SHEET = "個人班型摘要"
_SOLVER_SHEET = "求解與驗證資訊"
_PDF_SUMMARY_HEADER_OVERRIDES = {"週日出勤天數": "週日天數"}
_CJK_FONT = "ClinicScheduleCJK"
_CJK_FONT_BOLD = "ClinicScheduleCJKBold"
_CJK_FONT_ENVIRONMENT_VARIABLE = "CLINIC_SCHEDULER_PDF_FONT"
_CJK_BOLD_FONT_ENVIRONMENT_VARIABLE = "CLINIC_SCHEDULER_PDF_FONT_BOLD"
_BUNDLED_FONT_PACKAGE = "clinic_shift_scheduler.resources.fonts"
_BUNDLED_FONT_FILENAMES = (
    "NotoSansTC-Regular.ttf",
    "NotoSansTC-Bold.ttf",
)
_SYSTEM_CJK_FONT_PAIRS = (
    # Windows 10/11 Traditional Chinese default and broad system fallbacks.
    (Path("C:/Windows/Fonts/msjh.ttc"), Path("C:/Windows/Fonts/msjhbd.ttc")),
    (Path("C:/Windows/Fonts/msyh.ttc"), Path("C:/Windows/Fonts/msyhbd.ttc")),
    (Path("C:/Windows/Fonts/mingliu.ttc"), Path("C:/Windows/Fonts/mingliub.ttc")),
    # Development fallbacks for non-Windows hosts; v1 releases remain Windows-only.
    (
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"),
    ),
    (
        Path("/usr/share/fonts/truetype/noto/NotoSansTC-Regular.ttf"),
        Path("/usr/share/fonts/truetype/noto/NotoSansTC-Bold.ttf"),
    ),
    (Path("/System/Library/Fonts/PingFang.ttc"),) * 2,
)


def _try_register_font_pair(regular: Path, bold: Path) -> bool:
    if not regular.is_file() or not bold.is_file():
        return False
    try:
        regular_font = TTFont(_CJK_FONT, str(regular))
        bold_font = TTFont(_CJK_FONT_BOLD, str(bold))
    except Exception:
        return False
    pdfmetrics.registerFont(regular_font)
    pdfmetrics.registerFont(bold_font)
    return True


def _register_bundled_font_pair() -> bool:
    try:
        font_root = files(_BUNDLED_FONT_PACKAGE)
        regular_resource = font_root.joinpath(_BUNDLED_FONT_FILENAMES[0])
        bold_resource = font_root.joinpath(_BUNDLED_FONT_FILENAMES[1])
        if not regular_resource.is_file() or not bold_resource.is_file():
            return False
        with as_file(regular_resource) as regular, as_file(bold_resource) as bold:
            return _try_register_font_pair(regular, bold)
    except (FileNotFoundError, ModuleNotFoundError, TypeError):
        return False


def _register_cjk_font() -> None:
    registered = pdfmetrics.getRegisteredFontNames()
    if _CJK_FONT in registered and _CJK_FONT_BOLD in registered:
        return

    configured_regular = os.environ.get(_CJK_FONT_ENVIRONMENT_VARIABLE)
    if configured_regular:
        configured_bold = os.environ.get(
            _CJK_BOLD_FONT_ENVIRONMENT_VARIABLE,
            configured_regular,
        )
        if _try_register_font_pair(
            Path(configured_regular),
            Path(configured_bold),
        ):
            return

    if _register_bundled_font_pair():
        return

    for regular, bold in _SYSTEM_CJK_FONT_PAIRS:
        if _try_register_font_pair(regular, bold):
            return

    raise FormalExportError(
        "PDF export requires a Traditional Chinese TTF/TTC font; "
        "the bundled Noto Sans TC and system fallbacks were unavailable. "
        f"Set {_CJK_FONT_ENVIRONMENT_VARIABLE} and optionally "
        f"{_CJK_BOLD_FONT_ENVIRONMENT_VARIABLE} to valid font paths."
    )


def _key_value(sheet: Worksheet, label: str) -> Any | None:
    for row in sheet.iter_rows(min_col=1, max_col=2):
        if row[0].value == label:
            return row[1].value
    return None


def _validate_workbook(
    workbook: Any,
    *,
    provisional: bool = False,
) -> tuple[Worksheet, Worksheet]:
    expected_names = PROVISIONAL_WORKSHEET_NAMES if provisional else WORKSHEET_NAMES
    if tuple(workbook.sheetnames) != expected_names:
        raise FormalExportError(
            "PDF export requires the complete formal Excel workbook structure"
        )
    schedule = workbook[_SCHEDULE_SHEET]
    if schedule["A1"].value != "日期" or schedule["A2"].value != "星期":
        raise FormalExportError("PDF export requires a valid monthly schedule sheet")
    summary = workbook[_INDIVIDUAL_SUMMARY_SHEET]
    if summary["A1"].value != "姓名" or summary["E1"].value != "連續雙班日／出勤日":
        raise FormalExportError(
            "PDF export requires a valid individual pattern summary sheet"
        )
    solver = workbook[_SOLVER_SHEET]
    required_status = "FEASIBLE" if provisional else "OPTIMAL"
    if _key_value(solver, "正式狀態") != required_status:
        raise FormalExportError(
            f"PDF export requires Excel status {required_status}"
        )
    if _key_value(solver, "Validation") != "PASS":
        raise FormalExportError("PDF export requires Excel validation PASS")
    return schedule, summary


def _cell_text(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return f"{value.month}/{value.day}"
    return str(value)


def _fill_color(cell: Cell) -> colors.Color | None:
    fill = cell.fill
    if fill.fill_type != "solid":
        return None
    rgb = fill.fgColor.rgb
    if not isinstance(rgb, str) or len(rgb) not in (6, 8):
        return None
    return colors.HexColor(f"#{rgb[-6:]}")


def _font_color(cell: Cell) -> colors.Color | None:
    font_color = cell.font.color
    if font_color is None or font_color.type != "rgb":
        return None
    rgb = font_color.rgb
    if not isinstance(rgb, str) or len(rgb) not in (6, 8):
        return None
    return colors.HexColor(f"#{rgb[-6:]}")


def _paragraph(
    text: str,
    *,
    header: bool = False,
    text_color: colors.Color = colors.black,
) -> Paragraph:
    style = ParagraphStyle(
        name="schedule-header" if header else "schedule-cell",
        fontName=_CJK_FONT_BOLD,
        fontSize=7 if header else 6.2,
        leading=8 if header else 7.1,
        alignment=TA_CENTER,
        textColor=text_color,
        wordWrap="CJK",
    )
    escaped = (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    return Paragraph(escaped, style)


def _schedule_table(sheet: Worksheet) -> Table:
    values: list[list[Paragraph]] = []
    table_style: list[tuple[Any, ...]] = [
        ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#7890A0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, -1), _CJK_FONT_BOLD),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.2),
        ("TOPPADDING", (0, 0), (-1, -1), 2.2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2),
    ]
    for row_index, row in enumerate(
        sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        )
    ):
        rendered_row: list[Paragraph] = []
        for column_index, cell in enumerate(row):
            rendered_row.append(
                _paragraph(
                    _cell_text(cell),
                    header=row_index < 2 or column_index < 2,
                    text_color=_font_color(cell) or colors.black,
                )
            )
            background = _fill_color(cell)
            if background is not None:
                table_style.append(
                    (
                        "BACKGROUND",
                        (column_index, row_index),
                        (column_index, row_index),
                        background,
                    )
                )
        values.append(rendered_row)

    for merged_range in sheet.merged_cells.ranges:
        table_style.append(
            (
                "SPAN",
                (merged_range.min_col - 1, merged_range.min_row - 1),
                (merged_range.max_col - 1, merged_range.max_row - 1),
            )
        )

    usable_width = landscape(A4)[0] - 12 * mm
    period_width = 4.5 * mm
    role_width = 7.5 * mm
    date_width = (
        usable_width - period_width - role_width
    ) / max(1, sheet.max_column - 2)
    table = Table(
        values,
        colWidths=[
            period_width,
            role_width,
            *([date_width] * (sheet.max_column - 2)),
        ],
        repeatRows=2,
        hAlign="CENTER",
    )
    table.setStyle(TableStyle(table_style))
    return table


def _individual_summary_table(sheet: Worksheet) -> Table:
    values: list[list[Paragraph]] = []
    table_style: list[tuple[Any, ...]] = [
        ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#7890A0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, -1), _CJK_FONT_BOLD),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C5DFF0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
        ("TOPPADDING", (0, 0), (-1, -1), 1.8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.8),
    ]
    for row_index, row in enumerate(
        sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        )
    ):
        rendered_row: list[Paragraph] = []
        for column_index, cell in enumerate(row):
            text_color = colors.black
            if row_index > 0 and column_index == 0:
                text_color = _font_color(cell) or colors.black
            text = _cell_text(cell)
            if row_index == 0:
                text = _PDF_SUMMARY_HEADER_OVERRIDES.get(text, text)
            rendered_row.append(
                _paragraph(
                    text,
                    header=row_index == 0,
                    text_color=text_color,
                )
            )
        values.append(rendered_row)

    widths_mm = (11, 18, 12, 12, 36, 30, 36, 30, 18, 16)
    table = Table(
        values,
        colWidths=[width * mm for width in widths_mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    table.setStyle(TableStyle(table_style))
    return table


def _render_pdf(source: Path, target: Path, *, provisional: bool = False) -> None:
    workbook = load_workbook(source, read_only=False, data_only=True)
    try:
        schedule, summary = _validate_workbook(
            workbook,
            provisional=provisional,
        )
        title = workbook.properties.title or f"{source.stem} 月班表"
        _register_cjk_font()
        document = SimpleDocTemplate(
            str(target),
            pagesize=landscape(A4),
            leftMargin=6 * mm,
            rightMargin=6 * mm,
            topMargin=5 * mm,
            bottomMargin=5 * mm,
            title=f"{title}（本月班表）",
            author="clinic-shift-scheduler",
            subject=(
                "目前最佳合法班表（尚未完成最佳化）"
                if provisional
                else "正式月班表列印版"
            ),
        )
        title_style = ParagraphStyle(
            name="schedule-title",
            fontName=_CJK_FONT_BOLD,
            fontSize=11,
            leading=13,
            alignment=TA_CENTER,
            spaceAfter=0,
        )
        story = [Paragraph(f"{title}（本月班表）", title_style)]
        if provisional:
            warning_style = ParagraphStyle(
                name="provisional-warning",
                fontName=_CJK_FONT_BOLD,
                fontSize=8,
                leading=10,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#9A3412"),
                spaceAfter=0,
            )
            story.extend(
                (
                    Paragraph(
                        "尚未完成全部最佳化，不代表正式最佳結果",
                        warning_style,
                    ),
                    Spacer(1, 1 * mm),
                )
            )
        else:
            story.append(Spacer(1, 2 * mm))
        story.extend(
            (
                _schedule_table(schedule),
                Spacer(1, 3 * mm),
                _individual_summary_table(summary),
            )
        )
        document.build(story)
    finally:
        workbook.close()


def export_schedule_pdf_from_excel(
    excel_path: str | Path,
    *,
    output_path: str | Path | None = None,
    overwrite: bool = False,
) -> Path:
    """Create an atomic, single-sheet PDF using only the formal Excel workbook."""

    source = Path(excel_path)
    if not source.is_file():
        raise FileNotFoundError(f"formal Excel file not found: {source}")
    target = Path(output_path) if output_path is not None else source.with_suffix(".pdf")
    prepare_target(target, overwrite=overwrite)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{target.stem}.",
            suffix=".pdf",
            dir=target.parent,
            delete=False,
        ) as stream:
            temporary = Path(stream.name)
        _render_pdf(source, temporary)
        os.replace(temporary, target)
    finally:
        if temporary is not None and temporary.exists():
            temporary.unlink()
    return target


def export_provisional_schedule_pdf_from_excel(
    excel_path: str | Path,
    *,
    output_path: str | Path | None = None,
    overwrite: bool = False,
) -> Path:
    source = Path(excel_path)
    if not source.is_file():
        raise FileNotFoundError(f"provisional Excel file not found: {source}")
    target = Path(output_path) if output_path is not None else source.with_suffix(".pdf")
    prepare_target(target, overwrite=overwrite)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{target.stem}.",
            suffix=".pdf",
            dir=target.parent,
            delete=False,
        ) as stream:
            temporary = Path(stream.name)
        _render_pdf(source, temporary, provisional=True)
        os.replace(temporary, target)
    finally:
        if temporary is not None and temporary.exists():
            temporary.unlink()
    return target
