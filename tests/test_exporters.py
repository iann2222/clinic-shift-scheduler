from __future__ import annotations

import json
import tempfile
import unittest
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

import clinic_shift_scheduler.exporters.excel_exporter as excel_exporter
from clinic_shift_scheduler import (
    ExportFileExistsError,
    FeasibilityStatus,
    FormalExportError,
    RESULT_CONTRACT_NAME,
    RESULT_CONTRACT_VERSION,
    WORKSHEET_NAMES,
    build_output_paths,
    build_result_document,
    build_workbook,
    export_result_excel,
    export_result_json,
    export_schedule_pdf_from_excel,
    finalize_schedule_output,
    solve_lexicographic,
    validate_and_normalize,
)

from tests.fixtures import minimal_valid_input


class JsonExporterTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.data = validate_and_normalize(minimal_valid_input())
        cls.result = solve_lexicographic(cls.data)
        cls.output = finalize_schedule_output(cls.data, cls.result)
        assert cls.output.status is FeasibilityStatus.OPTIMAL

    def test_output_paths_use_stable_month_and_contract_version(self) -> None:
        paths = build_output_paths(self.data)

        self.assertEqual(paths.directory, Path("output"))
        self.assertEqual(paths.stem, "排班結果_2024-10.result-v1")
        self.assertEqual(paths.json, Path("output/排班結果_2024-10.result-v1.json"))
        self.assertEqual(paths.excel, Path("output/排班結果_2024-10.result-v1.xlsx"))
        self.assertEqual(paths.pdf, Path("output/排班結果_2024-10.result-v1.pdf"))

    def test_result_document_contains_the_versioned_formal_contract(self) -> None:
        generated_at = datetime(2024, 10, 2, 3, 4, 5, tzinfo=UTC)
        document = build_result_document(
            self.data,
            self.output,
            generated_at=generated_at,
        )

        self.assertEqual(
            document["contract"],
            {"name": RESULT_CONTRACT_NAME, "version": RESULT_CONTRACT_VERSION},
        )
        self.assertEqual(RESULT_CONTRACT_VERSION, "1.11")
        self.assertEqual(document["input_schema_version"], "v1")
        self.assertEqual(document["generated_at"], "2024-10-02T03:04:05Z")
        self.assertEqual(document["month"], "2024-10")
        self.assertEqual(document["status"], "OPTIMAL")
        self.assertIsNone(document["execution_timing"])
        self.assertIsNotNone(document["optimization_telemetry"])
        self.assertGreater(
            document["optimization_telemetry"]["assignment_variables"], 0
        )
        self.assertEqual(document["validation"]["status"], "PASS")
        self.assertNotIn("recomputed", document["validation"])
        self.assertEqual(
            document["objective_vector"],
            document["statistics"]["overall"]["objective_vector"],
        )
        self.assertEqual(len(document["stage_records"]), 16)
        self.assertIn("best_objective_bound", document["stage_records"][0])
        self.assertIn("num_conflicts", document["stage_records"][0])
        self.assertIn("num_branches", document["stage_records"][0])
        self.assertEqual(len(document["preference_benchmarks"]), 4)
        self.assertTrue(
            all(
                item["locked_actual_value"] is not None
                for item in document["preference_benchmarks"]
            )
        )
        self.assertEqual(len(document["class_pattern_locks"]), 2)
        self.assertEqual(len(document["assignments"]), 5)
        self.assertIn("individual", document["statistics"])
        self.assertIn("fairness_groups", document["statistics"])
        self.assertIn("class_preferences", document["statistics"])
        self.assertEqual(len(document["statistics"]["class_preferences"]), 4)
        self.assertNotIn("full_time_class_quality", document["statistics"])
        self.assertIn(
            "full_time_preference_rank1_max_regret",
            document["objective_vector"],
        )
        self.assertIn(
            "full_time_first_preference_ratio_total_gap",
            document["objective_vector"],
        )
        full_time_group = next(
            item
            for item in document["statistics"]["fairness_groups"]
            if item["employment_type"] == "full_time"
        )
        self.assertIn("ratio_basis_points", full_time_group)
        self.assertIn("ratio_gaps_basis_points", full_time_group)
        self.assertIn("rows", document["monthly_schedule"])

    def test_naive_generation_timestamp_is_rejected(self) -> None:
        with self.assertRaises(ValueError):
            build_result_document(
                self.data,
                self.output,
                generated_at=datetime(2024, 10, 2, 3, 4, 5),
            )

    def test_invalid_or_incomplete_result_cannot_be_formally_exported(self) -> None:
        invalid_output = replace(
            self.output,
            status=FeasibilityStatus.VALIDATION_FAILED,
            monthly_schedule=None,
        )

        with self.assertRaises(FormalExportError):
            build_result_document(self.data, invalid_output)

    def test_json_write_is_utf8_atomic_and_refuses_implicit_overwrite(self) -> None:
        generated_at = datetime(2024, 10, 2, 3, 4, 5, tzinfo=UTC)
        with tempfile.TemporaryDirectory() as directory:
            target = export_result_json(
                self.data,
                self.output,
                output_directory=directory,
                generated_at=generated_at,
            )
            document = json.loads(target.read_text(encoding="utf-8"))

            self.assertEqual(target.name, "排班結果_2024-10.result-v1.json")
            self.assertEqual(document["status"], "OPTIMAL")
            self.assertTrue(target.read_bytes().endswith(b"\n"))
            self.assertEqual(list(Path(directory).glob("*.tmp")), [])
            with self.assertRaises(ExportFileExistsError):
                export_result_json(
                    self.data,
                    self.output,
                    output_directory=directory,
                    generated_at=generated_at,
                )

            replaced = export_result_json(
                self.data,
                self.output,
                output_directory=directory,
                overwrite=True,
                generated_at=generated_at,
            )
            self.assertEqual(replaced, target)


class ExcelExporterTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.data = validate_and_normalize(minimal_valid_input())
        cls.result = solve_lexicographic(cls.data)
        cls.output = finalize_schedule_output(cls.data, cls.result)
        assert cls.output.status is FeasibilityStatus.OPTIMAL

    def test_workbook_has_expected_sheets_and_horizontal_schedule(self) -> None:
        workbook = build_workbook(self.data, self.output)
        try:
            self.assertEqual(tuple(workbook.sheetnames), WORKSHEET_NAMES)
            sheet = workbook["月班表"]
            schedule = self.output.monthly_schedule
            self.assertEqual(sheet["A1"].value, "日期")
            self.assertEqual(sheet["A2"].value, "星期")
            self.assertEqual(sheet["C1"].value, schedule.dates[0])
            self.assertEqual(sheet["C2"].value, "二")
            self.assertEqual(sheet["A3"].value, "早")
            self.assertEqual(sheet["B3"].value, "櫃台")
            self.assertEqual(sheet["B4"].value, "一診")
            self.assertEqual(sheet["B5"].value, "二診")
            self.assertEqual(sheet["A6"].value, "午")
            self.assertEqual(sheet["A9"].value, "晚")
            self.assertEqual(sheet["C3"].value, schedule.rows[0].cells[0].display)
            self.assertEqual(sheet["C5"].value, "—")
            self.assertEqual(sheet.freeze_panes, "C3")
            self.assertEqual(sheet.column_dimensions["A"].width, 3)
            self.assertEqual(sheet.column_dimensions["B"].width, 5)
            self.assertEqual(sheet["C1"].number_format, "m/d")
            self.assertIsNone(sheet["A1"].fill.fill_type)
            self.assertIsNone(sheet["A2"].fill.fill_type)
            merged = {str(item) for item in sheet.merged_cells.ranges}
            self.assertTrue({"A1:B1", "A2:B2", "A3:A5", "A6:A8", "A9:A11"}.issubset(merged))

            employee_colors: dict[str, str] = {}
            for row in range(3, 12):
                cell = sheet.cell(row, 3)
                if cell.value not in (None, "—", "休診"):
                    color = cell.font.color
                    self.assertIsNotNone(color)
                    assert color is not None
                    rgb = color.rgb[-6:]
                    if cell.value in employee_colors:
                        self.assertEqual(employee_colors[cell.value], rgb)
                    employee_colors[cell.value] = rgb
            self.assertEqual(
                len(set(employee_colors.values())),
                len(employee_colors),
            )
        finally:
            workbook.close()

    def test_statistics_and_solver_sheets_use_formal_output_values(self) -> None:
        workbook = build_workbook(self.data, self.output)
        try:
            summary = workbook["個人班型摘要"]
            summary_headers = {
                cell.value: cell.column
                for cell in summary[1]
                if cell.value is not None
            }
            first = self.output.individual_statistics[0]
            self.assertEqual(
                summary.cell(2, summary_headers["姓名"]).value,
                first.name,
            )
            self.assertEqual(
                summary.cell(2, summary_headers["總班次"]).value,
                first.total_shifts,
            )
            self.assertEqual(
                summary.cell(
                    2,
                    summary_headers["連續雙班日／出勤日"],
                ).value,
                f"{first.consecutive_double_days} / {first.attendance_days}"
                f"（{first.ratios['consecutive_double_days'].value:.1%}）",
            )
            self.assertEqual(
                summary.cell(
                    2,
                    summary_headers["週日節數"],
                ).value,
                first.sunday_shifts,
            )
            self.assertEqual(
                summary.cell(
                    2,
                    summary_headers["週日出勤天數"],
                ).value,
                first.sunday_attendance_days,
            )
            expected_sunday_dates = {
                assignment.date
                for assignment in self.output.assignments
                if assignment.employee_id == first.employee_id
                and assignment.date.weekday() == 6
            }
            self.assertEqual(
                first.sunday_attendance_days,
                len(expected_sunday_dates),
            )
            summary_name_colors = [
                summary.cell(row, 1).font.color.rgb[-6:]
                for row in range(2, summary.max_row + 1)
            ]
            self.assertEqual(
                len(set(summary_name_colors)),
                len(summary_name_colors),
            )

            individual = workbook["個人詳細統計"]
            headers = {
                cell.value: cell.column
                for cell in individual[1]
                if cell.value is not None
            }
            self.assertEqual(individual.cell(2, headers["employee_id"]).value, first.employee_id)
            self.assertEqual(individual.cell(2, headers["姓名"]).value, first.name)
            self.assertEqual(individual.cell(2, headers["總班次"]).value, first.total_shifts)
            self.assertGreaterEqual(individual.max_column, 30)
            for row in range(2, individual.max_row + 1):
                self.assertEqual(
                    individual.cell(row, headers["姓名"]).font.color.rgb[-6:],
                    summary.cell(row, 1).font.color.rgb[-6:],
                )

            groups = workbook["類別與公平性統計"]
            self.assertEqual(groups["A1"].value, "A／B 類別偏好理想值與公平退讓")
            self.assertEqual(groups["A2"].value, "正職類別")
            self.assertTrue(
                any(
                    groups.cell(row, 1).value == "類別統計"
                    for row in range(1, groups.max_row + 1)
                )
            )
            self.assertTrue(
                any(
                    groups.cell(row, 1).value == "A／B 類別偏好理想值與公平退讓"
                    for row in range(1, groups.max_row + 1)
                )
            )
            self.assertTrue(
                any(
                    groups.cell(row, 10).value == "退讓比例 (bp)"
                    for row in range(1, groups.max_row + 1)
                )
            )
            self.assertTrue(
                any(
                    groups.cell(row, 1).value == "Fairness group 統計"
                    for row in range(1, groups.max_row + 1)
                )
            )

            solver = workbook["求解與驗證資訊"]
            self.assertEqual(solver["A1"].value, "正式結果")
            self.assertEqual(solver["B4"].value, self.output.status.value)
            values = {
                solver.cell(row, 1).value: solver.cell(row, 2).value
                for row in range(1, solver.max_row + 1)
            }
            self.assertEqual(values["Validation"], "PASS")
            self.assertTrue(
                any(
                    solver.cell(row, 1).value == "類別偏好理想值 benchmark"
                    for row in range(1, solver.max_row + 1)
                )
            )
            self.assertTrue(
                any(
                    solver.cell(row, 1).value == "剩餘班型類別鎖"
                    for row in range(1, solver.max_row + 1)
                )
            )
            self.assertEqual(
                values["part_time_target_max_regret"],
                self.output.overall_statistics.objective_vector[
                    "part_time_target_max_regret"
                ],
            )
        finally:
            workbook.close()

    def test_pattern_summary_uses_validated_ratios_and_handles_no_attendance(self) -> None:
        workbook = build_workbook(self.data, self.output)
        try:
            sheet = workbook["個人班型摘要"]
            headers = {
                cell.value: cell.column
                for cell in sheet[1]
                if cell.value is not None
            }
            for row, stats in enumerate(self.output.individual_statistics, start=2):
                ratio = stats.ratios["single_shift_days"]
                percentage = "N/A" if ratio.value is None else f"{ratio.value:.1%}"
                self.assertEqual(
                    sheet.cell(row, headers["單節日／出勤日"]).value,
                    f"{ratio.numerator} / {ratio.denominator}（{percentage}）",
                )
            no_attendance = next(
                (stats for stats in self.output.individual_statistics if stats.attendance_days == 0),
                None,
            )
            if no_attendance is not None:
                row = self.output.individual_statistics.index(no_attendance) + 2
                self.assertEqual(
                    sheet.cell(row, headers["單節日／出勤日"]).value,
                    "0 / 0（N/A）",
                )
                self.assertEqual(sheet.cell(row, headers["週日節數"]).value, 0)
                self.assertEqual(
                    sheet.cell(row, headers["週日出勤天數"]).value,
                    0,
                )
        finally:
            workbook.close()

    def test_employee_colors_follow_stable_id_order_not_names(self) -> None:
        renamed = replace(
            self.output,
            individual_statistics=tuple(
                replace(stats, name=name)
                for stats, name in zip(
                    self.output.individual_statistics,
                    ("任意姓名三", "任意姓名一", "任意姓名二"),
                    strict=True,
                )
            ),
        )

        colors = excel_exporter._employee_color_map(renamed)

        self.assertEqual(colors["FT001"], "2860AD")
        self.assertEqual(colors["FT002"], "AD2833")
        self.assertEqual(colors["PT001"], "28AD60")

    def test_closed_date_cells_are_visually_distinct(self) -> None:
        payload = minimal_valid_input()
        payload["period"]["end_date"] = "2024-10-02"
        payload["period"]["closed_dates"] = ["2024-10-02"]
        data = validate_and_normalize(payload)
        output = finalize_schedule_output(data, solve_lexicographic(data))
        workbook = build_workbook(data, output)
        try:
            sheet = workbook["月班表"]
            self.assertEqual(sheet["D3"].value, "休診")
            self.assertEqual(sheet["D3"].fill.fgColor.rgb[-6:], "C9C9C9")
            self.assertNotEqual(
                sheet["C3"].fill.fgColor.rgb,
                sheet["D3"].fill.fgColor.rgb,
            )
        finally:
            workbook.close()

    def test_excel_write_reopens_and_refuses_implicit_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            target = export_result_excel(
                self.data,
                self.output,
                output_directory=directory,
            )
            reopened = load_workbook(target, read_only=False, data_only=False)
            try:
                self.assertEqual(tuple(reopened.sheetnames), WORKSHEET_NAMES)
                self.assertEqual(reopened["月班表"]["A1"].value, "日期")
                self.assertEqual(reopened["月班表"].freeze_panes, "C3")
            finally:
                reopened.close()
            self.assertEqual(target.name, "排班結果_2024-10.result-v1.xlsx")
            self.assertEqual(list(Path(directory).glob("*.tmp")), [])
            with self.assertRaises(ExportFileExistsError):
                export_result_excel(
                    self.data,
                    self.output,
                    output_directory=directory,
                )
            self.assertEqual(
                export_result_excel(
                    self.data,
                    self.output,
                    output_directory=directory,
                    overwrite=True,
                ),
                target,
            )

    def test_invalid_result_cannot_be_exported_to_excel(self) -> None:
        invalid_output = replace(
            self.output,
            status=FeasibilityStatus.VALIDATION_FAILED,
            monthly_schedule=None,
        )
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaises(FormalExportError):
                export_result_excel(
                    self.data,
                    invalid_output,
                    output_directory=directory,
                )
            self.assertEqual(list(Path(directory).iterdir()), [])


class PdfExporterTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.data = validate_and_normalize(minimal_valid_input())
        cls.result = solve_lexicographic(cls.data)
        cls.output = finalize_schedule_output(cls.data, cls.result)
        assert cls.output.status is FeasibilityStatus.OPTIMAL

    def test_pdf_is_single_page_monthly_schedule_derived_from_excel(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            excel = export_result_excel(
                self.data,
                self.output,
                output_directory=directory,
            )
            target = export_schedule_pdf_from_excel(excel)
            reader = PdfReader(target)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)

            self.assertEqual(target.name, "排班結果_2024-10.result-v1.pdf")
            self.assertEqual(len(reader.pages), 1)
            self.assertIn("月班表", text)
            self.assertIn("日期", text)
            self.assertIn("一診", text)
            self.assertIn("二診", text)
            self.assertIn("連續雙班日／出勤日", text)
            self.assertIn("單節日／出勤日", text)
            self.assertIn("週日節數", text)
            self.assertIn("週日天數", text)
            self.assertNotIn("週日出勤天數", text)
            self.assertIn(self.output.individual_statistics[0].name, text)
            self.assertNotIn("個人班型摘要", text)
            self.assertNotIn("Objective vector", text)
            self.assertEqual(list(Path(directory).glob("*.tmp")), [])

    def test_pdf_refuses_implicit_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            excel = export_result_excel(
                self.data,
                self.output,
                output_directory=directory,
            )
            target = export_schedule_pdf_from_excel(excel)
            with self.assertRaises(ExportFileExistsError):
                export_schedule_pdf_from_excel(excel)
            self.assertEqual(
                export_schedule_pdf_from_excel(excel, overwrite=True),
                target,
            )

    def test_pdf_rejects_excel_without_formal_pass_status(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            excel = export_result_excel(
                self.data,
                self.output,
                output_directory=directory,
            )
            workbook = load_workbook(excel)
            try:
                solver = workbook["求解與驗證資訊"]
                solver["B4"] = "FEASIBLE"
                workbook.save(excel)
            finally:
                workbook.close()
            with self.assertRaises(FormalExportError):
                export_schedule_pdf_from_excel(excel)


if __name__ == "__main__":
    unittest.main()
