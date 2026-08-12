from __future__ import annotations

import json
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from io import StringIO
from pathlib import Path
from time import sleep
from unittest.mock import patch

from openpyxl import load_workbook

from clinic_shift_scheduler import (
    CandidateDiagnosticSettings,
    CandidateExportConfig,
    CancellationToken,
    DiagnosticTimeSettings,
    EquivalentSolutionDiagnosticConfig,
    EquivalentSolutionDiagnosticStatus,
    ExecutionPhase,
    FeasibilityStatus,
    ProgressEvent,
    ProgressEventKind,
    RESULT_CONTRACT_VERSION,
    SchedulerAppConfig,
    ScheduleRunError,
    run_schedule_file,
)
import clinic_shift_scheduler.cli as cli_module
import clinic_shift_scheduler.console_app as console_module
from clinic_shift_scheduler.cli import main
from clinic_shift_scheduler.time_formatting import format_seconds_with_minutes
import clinic_shift_scheduler.runner as runner_module
import run_scheduler

from tests.fixtures import minimal_valid_input, synthetic_schedule_input


def progress_event(
    message: str,
    kind: ProgressEventKind = ProgressEventKind.INFORMATION,
) -> ProgressEvent:
    return ProgressEvent(
        phase=ExecutionPhase.APPLICATION,
        kind=kind,
        message=message,
    )


class ScheduleRunnerTests(unittest.TestCase):
    def test_cancelled_request_stops_before_reading_input(self) -> None:
        cancellation = CancellationToken()
        cancellation.cancel()

        with self.assertRaises(ScheduleRunError) as raised:
            run_schedule_file("missing.json", cancellation=cancellation)

        self.assertTrue(raised.exception.cancelled)
        self.assertEqual(
            raised.exception.issues[0].code,
            "operation_cancelled",
        )

    def test_interactive_console_overwrites_heartbeat_line(self) -> None:
        class InteractiveBuffer(StringIO):
            def isatty(self) -> bool:
                return True

        stream = InteractiveBuffer()
        printer = console_module.ConsoleProgressPrinter("[排班]", stream)

        printer(progress_event("嚴格分階段最佳化進行中：已耗時 5 秒", ProgressEventKind.HEARTBEAT))
        printer(progress_event("嚴格分階段最佳化進行中：已耗時 10 秒", ProgressEventKind.HEARTBEAT))
        heartbeat_output = stream.getvalue()
        self.assertNotIn("\n", heartbeat_output)
        self.assertEqual(heartbeat_output.count("\r"), 2)

        printer(progress_event("嚴格分階段最佳化完成：共耗時 20.1 秒"))
        final_output = stream.getvalue()
        self.assertIn("\r", final_output)
        self.assertTrue(final_output.endswith("20.1 秒\n"))

    def test_noninteractive_console_keeps_heartbeat_lines(self) -> None:
        stream = StringIO()
        printer = console_module.ConsoleProgressPrinter("[排班]", stream)

        printer(progress_event("嚴格分階段最佳化進行中：已耗時 5 秒", ProgressEventKind.HEARTBEAT))
        printer(progress_event("嚴格分階段最佳化進行中：已耗時 10 秒", ProgressEventKind.HEARTBEAT))

        self.assertEqual(stream.getvalue().count("\n"), 2)
        self.assertNotIn("\r", stream.getvalue())

    def test_interactive_console_overwrites_candidate_count(self) -> None:
        class InteractiveBuffer(StringIO):
            def isatty(self) -> bool:
                return True

        stream = InteractiveBuffer()
        printer = console_module.ConsoleProgressPrinter("[候選處理]", stream)

        printer(progress_event("已找到 1 份同品質候選班表", ProgressEventKind.CANDIDATE_COUNT))
        printer(progress_event("已找到 2 份同品質候選班表", ProgressEventKind.CANDIDATE_COUNT))

        self.assertNotIn("\n", stream.getvalue())
        self.assertEqual(stream.getvalue().count("\r"), 2)

    def test_noninteractive_console_keeps_candidate_count_lines(self) -> None:
        stream = StringIO()
        printer = console_module.ConsoleProgressPrinter("[候選處理]", stream)

        printer(progress_event("已找到 1 份同品質候選班表", ProgressEventKind.CANDIDATE_COUNT))
        printer(progress_event("已找到 2 份同品質候選班表", ProgressEventKind.CANDIDATE_COUNT))

        self.assertEqual(stream.getvalue().count("\n"), 2)
        self.assertNotIn("\r", stream.getvalue())

    def test_cli_defaults_follow_typed_application_defaults(self) -> None:
        defaults = SchedulerAppConfig(input_file="schedule.json")
        args = cli_module._parser().parse_args(["schedule.json"])

        self.assertEqual(
            args.equivalent_limit,
            defaults.candidate_diagnostic.search_limit,
        )
        self.assertEqual(
            args.equivalent_time_ratio,
            defaults.candidate_diagnostic.time.scheduling_time_ratio,
        )
        self.assertEqual(
            args.candidate_export_count,
            defaults.candidate_diagnostic.export_count,
        )
        self.assertEqual(
            args.candidate_export_formats,
            defaults.candidate_diagnostic.export_formats,
        )
        self.assertEqual(
            args.progress_interval,
            defaults.progress_update_seconds,
        )

    def test_elapsed_heartbeat_reports_progress_and_completion(self) -> None:
        events: list[ProgressEvent] = []

        def operation() -> str:
            sleep(0.035)
            return "done"

        result, elapsed = runner_module._run_with_elapsed_heartbeat(
            operation,
            events.append,
            interval_seconds=0.01,
        )

        self.assertEqual(result, "done")
        self.assertGreaterEqual(elapsed, 0.03)
        self.assertTrue(any(event.kind is ProgressEventKind.HEARTBEAT for event in events))
        self.assertIn("最佳化完成", events[-1].message)

    def test_total_elapsed_format_includes_minutes_without_sixty_seconds(self) -> None:
        self.assertEqual(
            format_seconds_with_minutes(185.34),
            "185.3 秒（約 3 分 5 秒）",
        )
        self.assertEqual(
            format_seconds_with_minutes(59.96),
            "60 秒（約 1 分 0 秒）",
        )

    def test_default_diagnostic_time_is_one_fifth_of_optimization(self) -> None:
        automatic = runner_module._resolve_diagnostic_config(
            EquivalentSolutionDiagnosticConfig(),
            125.0,
        )
        explicit = runner_module._resolve_diagnostic_config(
            EquivalentSolutionDiagnosticConfig(max_time_seconds=7.5),
            125.0,
        )
        proportional = runner_module._resolve_diagnostic_config(
            EquivalentSolutionDiagnosticConfig(scheduling_time_ratio=0.4),
            125.0,
        )

        self.assertEqual(automatic.max_time_seconds, 25.0)
        self.assertEqual(explicit.max_time_seconds, 7.5)
        self.assertEqual(proportional.max_time_seconds, 50.0)

    def test_primary_script_uses_configured_file_from_root_input(self) -> None:
        config = SchedulerAppConfig(
            input_file="排班輸入_2026-08.json",
            progress_update_seconds=7,
            candidate_diagnostic=CandidateDiagnosticSettings(
                search_limit=20,
                time=DiagnosticTimeSettings(
                    mode="定值",
                    fixed_seconds=12.5,
                    scheduling_time_ratio=None,
                ),
                export_count=2,
                export_formats=("json", "excel"),
            ),
        )
        with patch.object(
            run_scheduler,
            "load_scheduler_config",
            return_value=config,
        ), patch.object(
            run_scheduler,
            "run_schedule_request_with_console",
            return_value=0,
        ) as run_request:
            exit_code = run_scheduler.main()

        self.assertEqual(exit_code, 0)
        request = run_request.call_args.args[0]
        self.assertEqual(
            request.input_path,
            run_scheduler.PROJECT_ROOT
            / "input"
            / config.input_file,
        )
        self.assertEqual(
            request.output_directory,
            run_scheduler.PROJECT_ROOT / "output",
        )
        self.assertEqual(
            request.intermediate_directory,
            run_scheduler.PROJECT_ROOT / "runtime" / "expanded-input",
        )
        self.assertTrue(request.overwrite)
        self.assertEqual(request.progress_interval_seconds, 7)
        self.assertIsNotNone(request.diagnostic_config)
        assert request.diagnostic_config is not None
        self.assertEqual(request.diagnostic_config.max_alternatives, 20)
        self.assertEqual(
            request.diagnostic_config.max_time_seconds,
            12.5,
        )
        self.assertEqual(
            request.candidate_export_config.max_candidates,
            2,
        )

    def test_primary_script_reports_invalid_config_without_traceback(self) -> None:
        stderr = StringIO()
        with patch.object(
            run_scheduler,
            "load_scheduler_config",
            side_effect=ValueError("bad config"),
        ), redirect_stderr(stderr):
            exit_code = run_scheduler.main()

        self.assertEqual(exit_code, 1)
        self.assertIn("[設定] 無法讀取", stderr.getvalue())
        self.assertIn("bad config", stderr.getvalue())

    def test_complete_run_exports_files_and_records_execution_time(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            input_path = root / "input.json"
            payload = minimal_valid_input()
            payload["authoring_version"] = "weekly-v1"
            payload["period"] = {
                "start_date": "2024-10-01",
                "end_date": "2024-10-01",
                "holidays": ["2024-10-01"],
            }
            payload.pop("demands")
            payload["weekly_demands"] = [
                {
                    "weekdays": ["tuesday"],
                    "is_open": True,
                    "staffing": {
                        "morning": {"reception": 1, "assistant": 1},
                        "afternoon": {"reception": 1, "assistant": 1},
                        "evening": {"reception": 0, "assistant": 1},
                    },
                },
                {
                    "weekdays": [
                        "monday",
                        "wednesday",
                        "thursday",
                        "friday",
                        "saturday",
                        "sunday",
                    ],
                    "is_open": False,
                },
            ]
            payload["date_overrides"] = []
            input_path.write_text(
                json.dumps(payload, ensure_ascii=False),
                encoding="utf-8",
            )
            events: list[ProgressEvent] = []
            intermediate_directory = root / "runtime" / "expanded-input"
            intermediate_directory.mkdir(parents=True)
            (intermediate_directory / "obsolete.json").write_text(
                "{}",
                encoding="utf-8",
            )

            result = run_schedule_file(
                input_path,
                output_directory=root / "output",
                intermediate_directory=intermediate_directory,
                equivalent_solution_diagnostic_config=(
                    EquivalentSolutionDiagnosticConfig(
                        max_alternatives=1,
                        max_time_seconds=30,
                    )
                ),
                progress=events.append,
            )

            self.assertIs(result.output.status, FeasibilityStatus.OPTIMAL)
            self.assertTrue(result.precheck.status.value == "CONTINUE")
            self.assertTrue(result.json_path.is_file())
            self.assertTrue(result.excel_path.is_file())
            self.assertTrue(result.pdf_path.is_file())
            self.assertTrue(result.intermediate_input_path.is_file())
            self.assertFalse((intermediate_directory / "obsolete.json").exists())
            intermediate = json.loads(
                result.intermediate_input_path.read_text(encoding="utf-8")
            )
            self.assertNotIn("authoring_version", intermediate)
            self.assertNotIn("weekly_demands", intermediate)
            self.assertEqual(len(intermediate["demands"]), 6)
            self.assertGreaterEqual(result.total_execution_seconds, 0)
            self.assertGreaterEqual(result.json_export_seconds, 0)
            self.assertGreaterEqual(result.excel_export_seconds, 0)
            self.assertGreaterEqual(result.pdf_export_seconds, 0)
            self.assertGreaterEqual(result.formal_output_seconds, 0)
            self.assertIsNotNone(result.equivalent_solution_diagnostic)
            assert result.equivalent_solution_diagnostic is not None
            self.assertIn(
                result.equivalent_solution_diagnostic.status,
                (
                    EquivalentSolutionDiagnosticStatus.EXACT_COUNT,
                    EquivalentSolutionDiagnosticStatus.AT_LEAST_LIMIT,
                ),
            )
            self.assertGreaterEqual(
                result.equivalent_solution_diagnostic_seconds,
                0,
            )
            self.assertEqual(result.candidate_exports, ())
            messages = [event.message for event in events]
            self.assertTrue(any("最佳化" in message for message in messages))
            self.assertTrue(any("最佳化完成" in message for message in messages))
            self.assertTrue(
                any(
                    "完成：OPTIMAL + validation PASS" in message
                    for message in messages
                )
            )
            self.assertTrue(any("輸出檔案" in message for message in messages))
            combined_messages = "\n".join(messages)
            self.assertIn(
                "輸出正式檔案：JSON 保存完整結果，Excel 供查看與使用，"
                "PDF 由 Excel 月班表產生",
                combined_messages,
            )
            self.assertNotIn("輸出正式 JSON", combined_messages)
            self.assertNotIn("輸出正式 Excel", combined_messages)
            self.assertNotIn("由 Excel 月班表產生 PDF", combined_messages)
            self.assertIn(
                "輸出檔案（含 JSON、Excel、PDF）：",
                combined_messages,
            )
            self.assertIn(
                "[排班耗時] 完整排班時間（從讀檔到正式輸出）：",
                combined_messages,
            )
            self.assertNotIn("  正式 JSON：", combined_messages)
            self.assertNotIn("  正式 Excel：", combined_messages)
            self.assertNotIn("  月班表 PDF：", combined_messages)
            self.assertTrue(
                any(str(result.json_path) in message for message in messages)
            )

            timing = result.output.execution_timing
            self.assertIsNotNone(timing)
            assert timing is not None
            self.assertGreaterEqual(timing.input_loading_seconds, 0)
            self.assertGreaterEqual(timing.validation_normalization_seconds, 0)
            self.assertGreaterEqual(timing.precheck_seconds, 0)
            self.assertGreaterEqual(timing.optimization_seconds, 0)
            self.assertGreaterEqual(
                timing.result_validation_and_build_seconds,
                0,
            )
            self.assertGreaterEqual(timing.scheduling_pipeline_seconds, 0)

            document = json.loads(result.json_path.read_text(encoding="utf-8"))
            self.assertEqual(document["contract"]["version"], "1.9")
            self.assertEqual(RESULT_CONTRACT_VERSION, "1.9")
            self.assertAlmostEqual(
                document["execution_timing"]["optimization_seconds"],
                timing.optimization_seconds,
            )

            workbook = load_workbook(result.excel_path, read_only=True)
            try:
                sheet = workbook["求解與驗證資訊"]
                values = {
                    sheet.cell(row, 1).value: sheet.cell(row, 2).value
                    for row in range(1, sheet.max_row + 1)
                }
                self.assertAlmostEqual(
                    values["CP-SAT 最佳化（秒）"],
                    timing.optimization_seconds,
                )
                self.assertAlmostEqual(
                    values["排班管線總時間（秒）"],
                    timing.scheduling_pipeline_seconds,
                )
            finally:
                workbook.close()

    def test_exports_validated_candidate_json_and_clears_old_files(self) -> None:
        employees = [
            {
                "employee_id": employee_id,
                "name": employee_id,
                "employment_type": "full_time",
                "full_time_class": "A",
                "roles": ["assistant"],
                "fairness_group": "A_TEST",
                "shift_mode": "EXACT",
                "required_shifts": 1,
            }
            for employee_id in ("A1", "A2")
        ]
        payload = synthetic_schedule_input(
            start_date="2024-10-01",
            end_date="2024-10-01",
            roles=["assistant"],
            employees=employees,
            positive_demands={
                ("2024-10-01", "morning", "assistant"): 1,
                ("2024-10-01", "afternoon", "assistant"): 1,
            },
        )
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            input_path = root / "input.json"
            input_path.write_text(
                json.dumps(payload, ensure_ascii=False),
                encoding="utf-8",
            )
            candidate_directory = root / "output" / "候選班表"
            candidate_directory.mkdir(parents=True)
            stale = candidate_directory / "old.json"
            stale.write_text("{}", encoding="utf-8")

            result = run_schedule_file(
                input_path,
                output_directory=root / "output",
                intermediate_directory=root / "runtime" / "expanded-input",
                equivalent_solution_diagnostic_config=(
                    EquivalentSolutionDiagnosticConfig(
                        max_alternatives=1,
                        max_time_seconds=30,
                    )
                ),
                candidate_export_config=CandidateExportConfig(
                    max_candidates=1,
                    formats=("json", "excel", "pdf"),
                ),
            )

            self.assertFalse(stale.exists())
            self.assertEqual(len(result.candidate_exports), 1)
            candidate = result.candidate_exports[0]
            self.assertIsNotNone(candidate.json_path)
            assert candidate.json_path is not None
            self.assertTrue(candidate.json_path.is_file())
            self.assertIsNotNone(candidate.excel_path)
            self.assertIsNotNone(candidate.pdf_path)
            assert candidate.excel_path is not None
            assert candidate.pdf_path is not None
            self.assertTrue(candidate.excel_path.is_file())
            self.assertTrue(candidate.pdf_path.is_file())
            document = json.loads(candidate.json_path.read_text(encoding="utf-8"))
            self.assertEqual(document["status"], "OPTIMAL")
            self.assertEqual(document["validation"]["status"], "PASS")

            stdout = StringIO()

            def fake_run(request, callbacks):
                config = request.diagnostic_config
                assert config is not None
                self.assertIsNone(config.max_time_seconds)
                assert callbacks.progress is not None
                assert callbacks.diagnostic_progress is not None
                callbacks.progress(
                    progress_event(
                        "完成：OPTIMAL + validation PASS\n"
                        "  CP-SAT 最佳化：1.0 秒\n"
                        "[排班耗時] 完整排班時間（從讀檔到正式輸出）："
                        "2 秒（約 0 分 2 秒）\n"
                        "  輸出檔案：output/result.json"
                    )
                )
                callbacks.diagnostic_progress(
                    progress_event("開始搜尋同品質候選班表")
                )
                return result

            with patch(
                "clinic_shift_scheduler.console_app.run_schedule_application",
                side_effect=fake_run,
            ), redirect_stdout(stdout):
                exit_code = main([str(input_path)])
            self.assertEqual(exit_code, 0)
            printed = stdout.getvalue()
            self.assertIn("完成：OPTIMAL + validation PASS", printed)
            self.assertIn("CP-SAT 最佳化", printed)
            self.assertIn("從讀檔到正式輸出", printed)
            self.assertIn("[候選處理] 開始搜尋", printed)
            self.assertNotIn("[排班] 開始搜尋", printed)
            self.assertIn(
                "[執行] 總耗時（含完整排班與候選處理）：",
                printed,
            )


if __name__ == "__main__":
    unittest.main()
